#!/usr/bin/env python3
"""
XLSX to text converter using openpyxl with image extraction support.
Enhanced version for better text extraction from Excel files.
"""

import io
import logging
import os
import tempfile
import warnings
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

from ..core import BaseConverter, ConversionResult, ExtractedImage
from ..utils.ocr_client import is_skip_response

logger = logging.getLogger(__name__)


def _get_parallel_workers() -> int:
    """Get number of parallel workers for Vision API calls."""
    try:
        return int(os.getenv("VISION_API_PARALLEL_WORKERS", "200"))
    except ValueError:
        return 50


def _is_image_extraction_enabled() -> bool:
    """Check if image extraction is enabled via environment variable."""
    return os.getenv("ENABLE_IMAGE_EXTRACTION", "true").lower() in ("true", "1", "yes")


def _is_image_ai_description_enabled() -> bool:
    """Check if AI image description is enabled via environment variable."""
    return os.getenv("ENABLE_IMAGE_AI_DESCRIPTION", "true").lower() in ("true", "1", "yes")


def _get_image_max_width() -> int:
    """Get maximum image width for resizing."""
    try:
        return int(os.getenv("IMAGE_MAX_WIDTH", "1024"))
    except ValueError:
        return 1024


class XLSXConverter(BaseConverter):
    """Convert XLSX files to text using openpyxl with image extraction.
    
    Environment Variables:
        ENABLE_IMAGE_EXTRACTION: Enable/disable image extraction (default: true)
        ENABLE_IMAGE_AI_DESCRIPTION: Enable/disable Vision API descriptions (default: true)
        IMAGE_MAX_WIDTH: Maximum image width in pixels (default: 1024)
    """

    def __init__(self, max_rows: int = 1000):
        """Initialize XLSX converter.

        Args:
            max_rows: Maximum rows to process per sheet (for performance)
        """
        super().__init__()
        self.logger = logging.getLogger(__name__)
        self.max_rows = max_rows
        self._ocr_client = None
        # Set Japanese locale
        os.environ.update({
            'LC_ALL': 'ja_JP.UTF-8',
            'LC_CTYPE': 'ja_JP.UTF-8',
            'LANG': 'ja_JP.UTF-8'
        })

    def _get_ocr_client(self):
        """Lazy-load OCR client for Vision API."""
        if self._ocr_client is None and _is_image_ai_description_enabled():
            try:
                from ..utils import OCRClient
                self._ocr_client = OCRClient()
            except Exception as e:
                self.logger.warning(f"Failed to initialize OCR client: {e}")
        return self._ocr_client

    def _is_supported_extension(self, file_path: Path) -> bool:
        """Check if file is an Excel file."""
        return file_path.suffix.lower() in ['.xlsx', '.xlsm', '.xltx', '.xltm']

    def convert(self, file_path: Path) -> ConversionResult:
        """
        Convert XLSX to text with image extraction.

        Args:
            file_path: Path to Excel file

        Returns:
            ConversionResult with text content and extracted images
        """
        file_path = Path(file_path).resolve()
        
        try:
            # Validate file
            if not file_path.exists():
                return ConversionResult(
                    success=False,
                    message=f"File not found: {file_path}",
                    original_path=file_path,
                    file_format="xlsx",
                )
            
            if not self._is_supported_extension(file_path):
                return ConversionResult(
                    success=False,
                    message=f"Unsupported file extension: {file_path.suffix}",
                    original_path=file_path,
                    file_format="xlsx",
                )

            # Load workbook (not read_only to access images)
            # Note: read_only=True doesn't allow access to images
            # Suppress openpyxl DrawingML warning (shapes/drawings not fully supported)
            with warnings.catch_warnings():
                warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
                wb = load_workbook(filename=file_path, data_only=True)

            text_parts = []
            failed_sheets = []
            all_images: List[ExtractedImage] = []

            # Process each sheet: extract text and images per sheet
            sheet_data: List[Dict] = []  # Store sheet name, text, images
            
            for sheet_name in wb.sheetnames:
                try:
                    ws = wb[sheet_name]

                    # Extract data from sheet
                    sheet_text = self._extract_sheet_text(ws, self.max_rows)

                    # Extract images from sheet if enabled (without AI descriptions)
                    sheet_images = []
                    if _is_image_extraction_enabled():
                        sheet_images = self._extract_sheet_images_without_ai(ws, sheet_name)
                        all_images.extend(sheet_images)
                    
                    sheet_data.append({
                        "name": sheet_name,
                        "text": sheet_text,
                        "images": sheet_images,
                    })

                except Exception as e:
                    failed_sheets.append(sheet_name)
                    self.logger.warning(f"Failed to process sheet '{sheet_name}': {e}")
                    continue

            # Close workbook
            wb.close()
            
            # Get AI descriptions in parallel for all images
            if all_images and _is_image_ai_description_enabled():
                self.logger.info(f"Getting AI descriptions for {len(all_images)} images in parallel...")
                self._get_descriptions_parallel(all_images)

            # Check if any content was extracted
            if not sheet_data and not all_images:
                return ConversionResult(
                    success=False,
                    message="No content found in Excel file",
                    original_path=file_path,
                    file_format="xlsx",
                )

            # Build markdown: text + images per sheet (PDF style)
            for sheet in sheet_data:
                text_parts.append(f"## シート: {sheet['name']}")
                
                if sheet["text"]:
                    text_parts.append(sheet["text"])
                else:
                    text_parts.append("*(空のシート)*")
                
                # Embed images right after sheet text
                if sheet["images"]:
                    for img in sheet["images"]:
                        img_md = f"\n![{img.filename}](../04_images/{img.filename})\n"
                        if img.ai_description:
                            # 全行を > で囲む
                            desc_lines = img.ai_description.split('\n')
                            quoted_desc = '\n'.join(f"> {line}" for line in desc_lines)
                            img_md += f"\n> **図の説明:**\n>\n{quoted_desc}\n"
                        text_parts.append(img_md)

            # Join all text parts
            extracted_text = "\n\n".join(text_parts)

            # Build message
            message = f"Extracted {len(wb.sheetnames)} sheets, {len(all_images)} images"
            if failed_sheets:
                message += f" (failed: {failed_sheets})"

            self.logger.info(f"Successfully extracted {len(extracted_text)} characters, {len(all_images)} images")
            
            return ConversionResult(
                success=True,
                text=extracted_text,
                message=message,
                images=all_images,
                original_path=file_path,
                file_format="xlsx",
            )

        except FileNotFoundError:
            error_msg = f"File not found: {file_path}"
            self.logger.error(error_msg)
            return ConversionResult(
                success=False,
                message=error_msg,
                original_path=file_path,
                file_format="xlsx",
            )

        except Exception as e:
            error_msg = f"Excel conversion error: {str(e)}"
            self.logger.error(error_msg)
            return ConversionResult(
                success=False,
                message=error_msg,
                original_path=file_path,
                file_format="xlsx",
            )

    def _extract_sheet_images_without_ai(self, worksheet, sheet_name: str) -> List[ExtractedImage]:
        """Extract images from a worksheet WITHOUT AI descriptions.
        
        AI descriptions will be added later in parallel for better performance.
        
        Args:
            worksheet: openpyxl worksheet object
            sheet_name: Name of the sheet
            
        Returns:
            List of extracted images (without AI descriptions)
        """
        images = []
        max_width = _get_image_max_width()
        
        try:
            # Access images through _images attribute
            if not hasattr(worksheet, '_images'):
                return images
            
            for img_index, img in enumerate(worksheet._images):
                try:
                    # Get image data
                    image_data = img._data
                    if callable(image_data):
                        image_data = image_data()
                    
                    if not image_data:
                        continue
                    
                    # Determine image format
                    # openpyxl Image objects have a format attribute or we can detect from bytes
                    image_ext = "png"  # Default
                    if hasattr(img, 'format'):
                        image_ext = img.format.lower() if img.format else "png"
                    elif image_data[:4] == b'\x89PNG':
                        image_ext = "png"
                    elif image_data[:2] == b'\xff\xd8':
                        image_ext = "jpg"
                    elif image_data[:4] == b'GIF8':
                        image_ext = "gif"
                    
                    # Get image dimensions
                    width, height = self._get_image_dimensions(image_data)
                    
                    # Skip very small images
                    if width and height and (width < 50 or height < 50):
                        continue
                    
                    # Resize if too large
                    if width and width > max_width:
                        image_data = self._resize_image(image_data, max_width, image_ext)
                        if height:
                            new_height = int(height * max_width / width)
                            height = new_height
                        width = max_width
                    
                    # Generate filename
                    safe_sheet_name = sheet_name.replace("/", "_").replace("\\", "_")[:20]
                    filename = f"{safe_sheet_name}_img_{img_index + 1:03d}.{image_ext}"
                    
                    # Get anchor cell reference if available
                    cell_ref = ""
                    if hasattr(img, 'anchor') and img.anchor:
                        try:
                            if hasattr(img.anchor, '_from'):
                                cell_ref = f" (Cell: {img.anchor._from.col}{img.anchor._from.row})"
                        except:
                            pass
                    
                    # Create image without AI description (will be filled in parallel later)
                    images.append(ExtractedImage(
                        data=image_data,
                        filename=filename,
                        page_or_sheet=f"Sheet: {sheet_name}{cell_ref}",
                        position=img_index,
                        ai_description=None,  # Will be filled in parallel
                        width=width,
                        height=height,
                        format=image_ext,
                    ))
                    
                except Exception as e:
                    self.logger.warning(f"Failed to extract image {img_index} from sheet {sheet_name}: {e}")
                    continue
                    
        except Exception as e:
            self.logger.error(f"Image extraction failed for sheet {sheet_name}: {e}")
        
        return images
    
    def _get_descriptions_parallel(self, images: List[ExtractedImage]) -> None:
        """Get AI descriptions for multiple images in parallel.
        
        Updates the ai_description field of each ExtractedImage in place.
        Skips images that are identified as logos/icons/decorations.
        
        Args:
            images: List of ExtractedImage objects to process
        """
        ocr_client = self._get_ocr_client()
        if ocr_client is None:
            return
        
        max_workers = _get_parallel_workers()
        self.logger.info(f"Processing {len(images)} images with {max_workers} parallel workers")
        
        def get_description_for_image(img: ExtractedImage) -> Tuple[ExtractedImage, Optional[str]]:
            """Worker function to get description for a single image."""
            try:
                description = self._get_image_description(img.data, img.format or "png")
                # SKIPレスポンスのチェック
                if description and is_skip_response(description):
                    self.logger.debug(f"Skipped image {img.filename} (logo/icon/decoration)")
                    return (img, None)
                return (img, description)
            except Exception as e:
                self.logger.warning(f"Failed to get description for {img.filename}: {e}")
                return (img, None)
        
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {executor.submit(get_description_for_image, img): img for img in images}
            
            completed = 0
            skipped = 0
            for future in as_completed(futures):
                completed += 1
                try:
                    img, description = future.result()
                    if description is None:
                        skipped += 1
                    # Update the image object in place
                    object.__setattr__(img, 'ai_description', description)
                    if completed % 5 == 0:
                        self.logger.info(f"AI descriptions: {completed}/{len(images)} completed")
                except Exception as e:
                    self.logger.warning(f"Error processing image: {e}")
        
        self.logger.info(f"AI descriptions: {len(images)}/{len(images)} completed ({skipped} skipped)")

    def _get_image_dimensions(self, image_bytes: bytes) -> tuple:
        """Get image dimensions.
        
        Args:
            image_bytes: Image data
            
        Returns:
            (width, height) tuple or (None, None) if unable to determine
        """
        try:
            from PIL import Image
            img = Image.open(io.BytesIO(image_bytes))
            return img.size
        except Exception:
            return None, None

    def _resize_image(self, image_bytes: bytes, max_width: int, image_ext: str) -> bytes:
        """Resize image if it exceeds max width.
        
        Args:
            image_bytes: Original image bytes
            max_width: Maximum width in pixels
            image_ext: Image format extension
            
        Returns:
            Resized image bytes (or original if resize fails)
        """
        try:
            from PIL import Image
            
            img = Image.open(io.BytesIO(image_bytes))
            
            if img.width > max_width:
                ratio = max_width / img.width
                new_height = int(img.height * ratio)
                img = img.resize((max_width, new_height), Image.Resampling.LANCZOS)
                
                output = io.BytesIO()
                save_format = image_ext.upper()
                if save_format == "JPG":
                    save_format = "JPEG"
                img.save(output, format=save_format)
                return output.getvalue()
            
            return image_bytes
            
        except ImportError:
            self.logger.warning("PIL not available for image resizing")
            return image_bytes
        except Exception as e:
            self.logger.warning(f"Image resize failed: {e}")
            return image_bytes

    def _get_image_description(self, image_bytes: bytes, image_ext: str) -> Optional[str]:
        """Get AI-generated description of an image.
        
        Args:
            image_bytes: Image data
            image_ext: Image format extension
            
        Returns:
            AI-generated description or None
        """
        ocr_client = self._get_ocr_client()
        if ocr_client is None:
            return None
        
        try:
            # Write image to temp file
            with tempfile.NamedTemporaryFile(suffix=f".{image_ext}", delete=False) as f:
                f.write(image_bytes)
                temp_path = Path(f.name)
            
            try:
                # Use OCR client with description prompt
                # デフォルトプロンプト（ocr_client.py で定義）を使用
                success, description, error = ocr_client.image_to_markdown(temp_path)
                
                if success and description:
                    # 切り詰めなし - ベクトル検索のため全文を保持
                    return description.strip()
                    
            finally:
                # Clean up temp file
                if temp_path.exists():
                    temp_path.unlink()
                    
        except Exception as e:
            self.logger.warning(f"Failed to get image description: {e}")
        
        return None

    def _extract_sheet_text(self, worksheet, max_rows: int) -> str:
        """
        Extract worksheet content as a Markdown table.

        The first non-empty row is treated as a header. All values are
        sanitised so Markdown pipes and embedded newlines do not break the
        table layout.
        
        結合セルに対応: 結合されたセルの値を結合範囲内のすべてのセルに反映します。
        """
        def _format_cell(value) -> str:
            if value is None:
                return ""
            if isinstance(value, bool):
                text = "TRUE" if value else "FALSE"
            else:
                text = str(value).strip()

            # Preserve line breaks inside cells for Markdown rendering
            text = text.replace("\r\n", "\n").replace("\r", "\n").replace("\n", "<br>")
            # Escape pipe characters so they do not split columns
            text = text.replace("|", "\\|")
            return text

        # 結合セルのマッピングを構築
        # 結合範囲内の全セルに左上セルの値をマッピング
        merged_cell_map: Dict[Tuple[int, int], any] = {}
        try:
            for merged_range in worksheet.merged_cells.ranges:
                # 結合範囲の左上セルの値を取得
                top_left_value = worksheet.cell(
                    merged_range.min_row, 
                    merged_range.min_col
                ).value
                
                # 結合範囲内の全セルにその値をマッピング
                for row_idx in range(merged_range.min_row, merged_range.max_row + 1):
                    for col_idx in range(merged_range.min_col, merged_range.max_col + 1):
                        merged_cell_map[(row_idx, col_idx)] = top_left_value
        except Exception as e:
            self.logger.warning(f"結合セル情報の取得に失敗: {e}")

        rows = []
        rows_to_process = min(worksheet.max_row, max_rows) if worksheet.max_row else max_rows

        for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=rows_to_process), start=1):
            formatted = []
            for col_idx, cell in enumerate(row, start=1):
                # 結合セルの場合はマッピングから値を取得
                if (row_idx, col_idx) in merged_cell_map:
                    value = merged_cell_map[(row_idx, col_idx)]
                else:
                    value = cell.value
                formatted.append(_format_cell(value))
            
            # Skip completely empty rows
            if any(cell for cell in formatted):
                rows.append(formatted)

        if not rows:
            return ""

        # Keep column count consistent across all rows
        max_cols = max(len(r) for r in rows)
        normalized_rows = [r + [""] * (max_cols - len(r)) for r in rows]

        header = normalized_rows[0]
        data_rows = normalized_rows[1:]

        header_line = "| " + " | ".join(header) + " |"
        separator_line = "| " + " | ".join(["---"] * max_cols) + " |"
        data_lines = ["| " + " | ".join(row) + " |" for row in data_rows] if data_rows else []

        table_lines = [header_line, separator_line, *data_lines]
        return "\n".join(table_lines)
